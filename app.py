import html
import io
import os
import re

import google.generativeai as genai
import pandas as pd
import plotly.express as px
import requests
import streamlit as st
from openpyxl import load_workbook

# --- 1. CONFIGURAÇÃO VISUAL E VARIÁVEIS GLOBAIS ---
st.set_page_config(page_title="DASHBOARD NNUP · SEJUSP MG", layout="wide", page_icon="📊")

# Dicionário de meses para o filtro
MAPA_MESES = {
    "Janeiro": "Jan", "Fevereiro": "Fev", "Março": "Mar", "Abril": "Abr", 
    "Maio": "Mai", "Junho": "Jun", "Julho": "Jul", "Agosto": "Ago", 
    "Setembro": "Set", "Outubro": "Out", "Novembro": "Nov", "Dezembro": "Dez"
}
LISTA_MESES = ["Todos"] + list(MAPA_MESES.keys())


def _secret(key: str, default: str = "") -> str:
    try:
        v = st.secrets.get(key, default)
        return str(v).strip() if v not in (None, "") else default
    except Exception:
        return os.environ.get(key, default)


def get_gemini_key() -> str:
    return _secret("GEMINI_API_KEY", "") or os.environ.get("GEMINI_API_KEY", "")


# --- CSS: tema escuro minimalista ---
st.markdown(
    """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    html, body, .stApp { font-family: 'Outfit', system-ui, sans-serif !important; }
    .main .block-container { padding-top: 1rem; max-width: 1180px; }
    header { visibility: hidden; height: 0; }

    [data-testid="stMetric"] {
        background: #151b22 !important;
        border: 1px solid #252d38;
        border-radius: 12px;
        padding: 0.85rem 1rem;
    }
    [data-testid="stMetricValue"] { font-size: 1.5rem !important; color: #5ec8e8 !important; font-weight: 600 !important; }
    [data-testid="stMetricLabel"] { font-size: 0.82rem !important; color: #8b9cad !important; text-transform: uppercase; letter-spacing: 0.06em; }

    /* Abas ÁGUA / ENERGIA — mesma largura e altura, rótulos em maiúsculas */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px !important;
        background: transparent;
        padding: 0 0 10px;
        border-bottom: 1px solid #252d38;
        display: flex !important;
        width: 100% !important;
    }
    .stTabs [data-baseweb="tab-list"] > [data-baseweb="tab"],
    .stTabs [data-baseweb="tab-list"] > li {
        flex: 1 1 0 !important;
        min-width: 0 !important;
        max-width: none !important;
        height: 48px !important;
        min-height: 48px !important;
        border-radius: 10px !important;
        border: none !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        color: #8b9cad !important;
        text-transform: uppercase !important;
        letter-spacing: 0.06em !important;
        justify-content: center !important;
        align-items: center !important;
    }
    .stTabs [data-baseweb="tab"]:hover { background: #1a222c !important; color: #e6edf3 !important; }
    .stTabs [aria-selected="true"] {
        background: #1e2a35 !important;
        border: 1px solid #2a3f4d !important;
    }
    .stTabs [aria-selected="true"] div { color: #7dd3f0 !important; }

    div[role="radiogroup"] { display: flex; gap: 8px; flex-wrap: wrap; }
    div[role="radiogroup"] > label {
        background: #151b22 !important;
        padding: 8px 14px !important;
        border-radius: 8px !important;
        border: 1px solid #252d38 !important;
    }
    div[role="radiogroup"] > label:hover { border-color: #3d5a6e !important; }
    div[role="radiogroup"] > label:has(input:checked) {
        background: #1e2a35 !important;
        border-color: #42a5c8 !important;
    }
    div[role="radiogroup"] > label:has(input:checked) p,
    div[role="radiogroup"] > label:has(input:checked) div { color: #b8e8ff !important; }

    .hero-banner {
        background: linear-gradient(145deg, #0f1620 0%, #15202c 45%, #0c1218 100%);
        border: 1px solid #2a3d4d;
        border-radius: 16px;
        padding: 1.35rem 1.6rem 1.3rem;
        margin-bottom: 1.25rem;
        box-shadow: 0 12px 40px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(255,255,255,0.04);
        text-align: center;
    }
    .hero-banner .hero-kicker {
        display: block;
        font-size: 0.7rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: #42a5c8 !important;
        font-weight: 700;
        margin: 0 auto 0.35rem;
    }
    .hero-banner .hero-main-title {
        font-size: clamp(2.1rem, 5.5vw, 3.75rem);
        font-weight: 800;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: #f0f4f8 !important;
        margin: 0 auto;
        line-height: 1.08;
        text-shadow: 0 2px 24px rgba(66, 165, 200, 0.15);
    }
    .hero-banner .hero-sub {
        color: #9eb0c2 !important;
        font-size: clamp(0.95rem, 1.6vw, 1.1rem);
        margin: 0.75rem auto 0;
        font-weight: 400;
        line-height: 1.45;
        max-width: 56ch;
    }
    .hero-banner .hero-accent {
        height: 3px;
        width: 72px;
        background: linear-gradient(90deg, #42a5c8, #5ec8e8, transparent);
        border-radius: 2px;
        margin: 1rem auto 0;
    }
    section.main [data-testid="stTabs"] { margin-top: 0.15rem; }

    .panel-insights {
        background: #12171d;
        border: 1px solid #222a33;
        border-radius: 12px;
        padding: 1rem 1.15rem;
        margin-bottom: 1rem;
    }
    .panel-insights h4 { margin-top: 0 !important; color: #e6edf3 !important; font-weight: 600; font-size: 1rem; }
    .answer-box {
        background: rgba(66, 165, 200, 0.07);
        border: 1px solid #2a3f4d;
        border-radius: 10px;
        padding: 0.9rem 1rem;
        margin-top: 0.65rem;
        color: #d1dae6 !important;
        line-height: 1.55;
        word-break: normal !important;
        overflow-wrap: break-word;
        white-space: normal !important;
        letter-spacing: normal !important;
    }

    .warning-box { background: #1f1a12; padding: 12px 14px; border-radius: 8px; border-left: 3px solid #c9a227; color: #e8d49a !important; margin-bottom: 8px; }
    .danger-box { background: #1f1215; padding: 12px 14px; border-radius: 8px; border-left: 3px solid #c94c5c; color: #f0c4ca !important; margin-bottom: 8px;}
    .money-box { background: #0f1a14; padding: 12px 14px; border-radius: 8px; border-left: 3px solid #3d9a6a; color: #b8e6cc !important; margin-bottom: 8px; }

    div[data-testid="stExpander"] { background: #151b22; border: 1px solid #252d38; border-radius: 10px; }
</style>
""",
    unsafe_allow_html=True,
)

# --- 2. FUNÇÕES DE SUPORTE E FORMATAÇÃO ---
def formatar_moeda_br(valor):
    if pd.isna(valor): return "R$ 0,00"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_numero_br(valor):
    if pd.isna(valor): return "0"
    return f"{valor:,.0f}".replace(",", ".")


def plotly_dark(fig):
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#9fb0c3"),
        title_font=dict(color="#e6edf3", size=14),
        margin=dict(t=48, b=40, l=48, r=24),
    )
    fig.update_xaxes(
        gridcolor="rgba(255,255,255,0.06)",
        zerolinecolor="rgba(255,255,255,0.08)",
        tickfont=dict(color="#8b9cad"),
    )
    fig.update_yaxes(
        gridcolor="rgba(255,255,255,0.06)",
        zerolinecolor="rgba(255,255,255,0.08)",
        tickfont=dict(color="#8b9cad"),
    )
    return fig


def escapar_texto_para_caixa_resposta(texto: str) -> str:
    """Texto puro → HTML seguro; evita markdown/HTML partido dentro de divs."""
    if texto is None:
        return ""
    t = str(texto).replace("\r\n", "\n").replace("\r", "\n")
    return html.escape(t, quote=False).replace("\n", "<br/>")


def render_caixa_resposta(texto: str) -> None:
    """Uma única chamada st.markdown: conteúdo não passa pelo parser de markdown."""
    if not (texto and str(texto).strip()):
        return
    inner = escapar_texto_para_caixa_resposta(texto)
    st.markdown(
        f'<div class="answer-box">{inner}</div>',
        unsafe_allow_html=True,
    )


def resposta_pergunta_rapida(df, utilidade, qid):
    if df.empty:
        return "Não há dados no filtro atual."
    col_c = "Volume_M3" if utilidade == "Água" else "Consumo_KWh"
    um = "m³" if utilidade == "Água" else "kWh"

    if qid == "q1":
        g = df.groupby("Unidade", as_index=False)["Valor_Total"].sum().sort_values("Valor_Total", ascending=False)
        if g.empty:
            return "Sem valores de custo."
        row = g.iloc[0]
        u = str(row["Unidade"]).strip()
        return f"Maior custo acumulado: {u} — {formatar_moeda_br(row['Valor_Total'])} no período filtrado."

    if qid == "q2":
        g = df.groupby("Unidade", as_index=False)[col_c].sum().sort_values(col_c, ascending=False)
        if g.empty or g.iloc[0][col_c] <= 0:
            return "Sem consumo consolidado para ranquear."
        row = g.iloc[0]
        u = str(row["Unidade"]).strip()
        return f"Maior consumo físico: {u} — {formatar_numero_br(row[col_c])} {um}."

    if qid == "q3":
        g = df.groupby(["Nome_Mes", "Mes"], as_index=False)["Valor_Total"].sum().sort_values("Valor_Total", ascending=False)
        if g.empty:
            return "Sem faturamento por mês."
        row = g.iloc[0]
        return f"Mês de pico de faturamento: {row['Nome_Mes']} — {formatar_moeda_br(row['Valor_Total'])}."

    if qid == "q4":
        g = df.groupby("Origem")["Valor_Total"].sum().sort_values(ascending=False)
        if g.empty:
            return "Sem origem (aba) no filtro."
        tot = g.sum()
        u, v = g.index[0], float(g.iloc[0])
        pct = 100 * v / tot if tot else 0
        return f"Origem que mais concentra custo: {u} — {formatar_moeda_br(v)} ({pct:.1f}% do total filtrado)."

    if qid == "q5":
        g = df.groupby("Unidade", as_index=False)["Valor_Total"].sum().sort_values("Valor_Total", ascending=False).head(3)
        if g.empty:
            return "Sem dados para top 3."
        lines = [
            f"{i}. {str(row['Unidade']).strip()} — {formatar_moeda_br(row['Valor_Total'])}"
            for i, row in enumerate(g.to_dict("records"), 1)
        ]
        return "Top 3 unidades por custo:\n" + "\n".join(lines)

    if qid == "q6":
        agg = df.groupby("Unidade").agg(Valor_Total=("Valor_Total", "sum"), Cons=(col_c, "sum")).reset_index()
        agg = agg[agg["Cons"] > 0]
        if agg.empty:
            return "Sem consumo para calcular tarifa média por unidade."
        agg["tarifa"] = agg["Valor_Total"] / agg["Cons"]
        row = agg.sort_values("tarifa", ascending=False).iloc[0]
        u = str(row["Unidade"]).strip()
        return (
            f"Maior tarifa média (R$ por {um}): {u} — {formatar_moeda_br(row['tarifa'])} por {um} "
            f"(custo total {formatar_moeda_br(row['Valor_Total'])})."
        )

    if qid == "q7":
        if "Var_Financeiro_Pct" not in df.columns or "Valor_Anterior" not in df.columns:
            return "Dados sem histórico mês a mês para esta análise."
        sub = df[(df["Valor_Total"] > 100) & (df["Valor_Anterior"].notna()) & (df["Valor_Anterior"] > 0)].copy()
        sub = sub.dropna(subset=["Var_Financeiro_Pct"])
        if sub.empty:
            return "Não há histórico suficiente para comparar variação financeira mês a mês."
        worst = sub.loc[sub["Var_Financeiro_Pct"].idxmax()]
        return (
            f"Maior alta financeira (mês sobre mês): {worst['Unidade']} em {worst['Nome_Mes']} — "
            f"variação {worst['Var_Financeiro_Pct']:.1f}% (valor {formatar_moeda_br(worst['Valor_Total'])})."
        )

    if qid == "q8":
        vt = df["Valor_Total"].sum()
        ve = df["Valor_Encargos"].sum() if "Valor_Encargos" in df.columns else 0
        if vt <= 0:
            return "Total faturado zerado no filtro."
        pe = 100 * ve / vt if vt else 0
        return f"Encargos e multas somam {formatar_moeda_br(ve)} ({pe:.1f}% do faturamento de {formatar_moeda_br(vt)})."

    return ""


def extrair_mes_do_nome(texto):
    t = texto.lower().strip()
    nomes = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 
             'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
    for nome, num in nomes.items():
        if nome in t: return num
    match_final = re.search(r'(\d+)$', t)
    if match_final:
        num = int(match_final.group(1))
        if 1 <= num <= 12: return num
    return None

def limpar_valor_universal(valor, tipo_dado):
    if valor is None: return 0
    if isinstance(valor, (int, float)): val_float = float(valor)
    else:
        s = str(valor).strip().upper()
        if s in ['', '-', '–', 'NAN', 'NONE', '#DIV/0!', '#REF!', '#VALUE!', 'R$ -']: return 0
        s_limpa = re.sub(r'[^\d,-]', '', s)
        if not s_limpa: return 0
        s_limpa = s_limpa.replace('.', '').replace(',', '.')
        try: val_float = float(s_limpa)
        except: return 0
    if tipo_dado in ['Volume_M3', 'Consumo_KWh']:
        if val_float < 0: return 0
        if val_float > 5000000: return 0
    return val_float

def detectar_linha_cabecalho(linhas):
    for i, row in enumerate(linhas[:10]): 
        txt = " ".join([str(c).upper() for c in row if c])
        if "UNIDADES PRISIONAIS" in txt or "UNIDADES DE INTEGRAÇÃO" in txt or "UNIDADE" in txt: return i
    return 0

# --- 3. CARREGAMENTO DE DADOS ---
def _abrir_planilha(origem):
    if isinstance(origem, str) and origem.startswith(("http://", "https://")):
        r = requests.get(origem, timeout=180, headers={"User-Agent": "DashboardDSG/1.0"})
        r.raise_for_status()
        return load_workbook(io.BytesIO(r.content), data_only=True)
    return load_workbook(origem, data_only=True)


def process_workbook(wb, utilidade="agua"):
    lista_dados = []
    try:
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            linhas = list(ws.values)
            if not linhas: continue
            
            idx_header = detectar_linha_cabecalho(linhas)
            header = linhas[idx_header]
            dados = linhas[idx_header+1:]
            
            idx_unidade = -1
            for i, col in enumerate(header):
                if not col: continue
                t = str(col).upper().strip()
                if t in ["UNIDADES PRISIONAIS", "UNIDADES DE INTEGRAÇÃO", "UNIDADE PRISIONAL"]:
                    idx_unidade = i
                    break
            if idx_unidade == -1: 
                for i, col in enumerate(header):
                    if col and "UNIDADE" in str(col).upper():
                        idx_unidade = i
                        break
            if idx_unidade == -1: idx_unidade = 0
            
            palavras_chave_cadastrais = [
                'UNIDADE EXECUTORA', 'UPG', 'HIDRÔMETRO', 'CONCESSIONÁRIA', 'CNPJ', 
                'IDENTIFICADOR', 'EMPENHO', 'MATRÍCULA', 'FONTE 10', 'FONTE 60'
            ]
            mapa_cadastral = {}
            for i, col in enumerate(header):
                if not col: continue
                col_nome = str(col).upper().strip()
                for pc in palavras_chave_cadastrais:
                    if pc in col_nome:
                        mapa_cadastral[col_nome] = i
                        break

            mapa = {m: {} for m in range(1, 13)}
            ultimo_mes = 1
            
            for idx, col_obj in enumerate(header):
                if not col_obj: continue
                col_nome = str(col_obj)
                t_lower = col_nome.lower()
                
                mes_encontrado = extrair_mes_do_nome(col_nome)
                if mes_encontrado: 
                    ultimo_mes = mes_encontrado
                    mes_uso = mes_encontrado
                else: 
                    mes_uso = ultimo_mes
                
                tipo = None
                if 'empenho' in t_lower or 'emp ' in t_lower or 'código' in t_lower or 'id ' in t_lower or 'fonte' in t_lower: continue 

                if ('m³' in t_lower or 'm3' in t_lower or 'kwh' in t_lower or 'consumo' in t_lower) and 'valor' not in t_lower:
                    proibidos = ['variação', 'variacao', 'diferença', 'financeiro', '%']
                    if not any(p in t_lower for p in proibidos): 
                        tipo = 'Volume_M3' if utilidade == 'agua' else 'Consumo_KWh'
                elif 'líquido' in t_lower or 'liquido' in t_lower or ('total' in t_lower and ('fatura' in t_lower or '(' in t_lower)): 
                    tipo = 'Valor_Total'
                elif t_lower.strip() == 'total': 
                    tipo = 'Valor_Total'
                elif ('encargo' in t_lower or 'multa' in t_lower): 
                    tipo = 'Valor_Encargos'
                
                if tipo:
                    if tipo == 'Valor_Total' and 'Valor_Total' in mapa[mes_uso]:
                        if 'líquido' in t_lower: mapa[mes_uso][tipo] = idx
                    else:
                        mapa[mes_uso][tipo] = idx

            for linha in dados:
                if idx_unidade >= len(linha): continue
                nome_unidade = linha[idx_unidade]
                if not nome_unidade: continue
                
                u_str = str(nome_unidade).strip().upper()
                if not u_str or len(u_str) < 3: continue
                if any(x in u_str for x in ['TOTAL', 'TOTAIS', 'GERAL', 'CNPJ', 'CONCESSIONÁRIA', 'NOME DA']): continue
                if "UNIDADE" in u_str and ("RESPONSÁVEL" in u_str or "PRISIONAL" in u_str or "DE INTEGRA" in u_str): continue
                if u_str.replace('.','').replace('-','').isdigit(): continue 
                
                info_cadastral = {}
                for nome_col, idx_col in mapa_cadastral.items():
                    if idx_col < len(linha) and linha[idx_col]:
                        info_cadastral[nome_col] = str(linha[idx_col]).strip()
                
                for m in range(1, 13):
                    cols_mes = mapa[m]
                    reg = {'Origem': nome_aba, 'Unidade': str(nome_unidade).strip(), 'Mes': m, 
                           'Valor_Total': 0, 'Valor_Encargos': 0}
                    if utilidade == 'agua': reg['Volume_M3'] = 0
                    else: reg['Consumo_KWh'] = 0

                    if cols_mes:
                        for k, idx_col in cols_mes.items():
                            if idx_col < len(linha):
                                reg[k] = limpar_valor_universal(linha[idx_col], k)
                    
                    reg.update(info_cadastral)
                    lista_dados.append(reg)

        return pd.DataFrame(lista_dados)
    except Exception as e:
        st.error(f"Erro na leitura de {utilidade.upper()}: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner="Carregando planilha…")
def carregar_dados_cache(fonte: str, utilidade: str):
    try:
        wb = _abrir_planilha(fonte)
        return process_workbook(wb, utilidade)
    except Exception as e:
        st.error(f"Não foi possível abrir a fonte ({utilidade}): {e}")
        return pd.DataFrame()


def carregar_dados_upload(arquivo, utilidade: str):
    try:
        wb = load_workbook(arquivo, data_only=True)
        return process_workbook(wb, utilidade)
    except Exception as e:
        st.error(f"Erro no arquivo enviado ({utilidade}): {e}")
        return pd.DataFrame()

def enriquecer_dados(df, utilidade='agua'):
    if df.empty: return df
    col_consumo = 'Volume_M3' if utilidade == 'agua' else 'Consumo_KWh'
    
    df = df.sort_values(by=['Unidade', 'Mes'])
    df['Preco_Medio'] = df.apply(lambda x: x['Valor_Total'] / x[col_consumo] if x[col_consumo] > 0 else 0, axis=1)
    df['Consumo_Anterior'] = df.groupby('Unidade')[col_consumo].shift(1)
    df['Valor_Anterior'] = df.groupby('Unidade')['Valor_Total'].shift(1)
    
    def calc_var(atual, anterior):
        if anterior and anterior > 0: return ((atual - anterior) / anterior) * 100
        return 0
    
    df['Var_Consumo_Pct'] = df.apply(lambda x: calc_var(x[col_consumo], x['Consumo_Anterior']), axis=1)
    df['Var_Financeiro_Pct'] = df.apply(lambda x: calc_var(x['Valor_Total'], x['Valor_Anterior']), axis=1)
    
    df['Valor_Total_Fmt'] = df['Valor_Total'].apply(formatar_moeda_br)
    df['Consumo_Fmt'] = df[col_consumo].apply(formatar_numero_br)
    
    nomes_mes_map = {1:'Jan', 2:'Fev', 3:'Mar', 4:'Abr', 5:'Mai', 6:'Jun', 7:'Jul', 8:'Ago', 9:'Set', 10:'Out', 11:'Nov', 12:'Dez'}
    df['Nome_Mes'] = df['Mes'].map(nomes_mes_map)
    return df

def exibir_top_15(df, utilidade):
    st.markdown("---")
    st.markdown(f"**Rankings · {utilidade}**")
    st.caption("Top 15 com base nos filtros atuais.")

    col_consumo = 'Volume_M3' if utilidade == 'Água' else 'Consumo_KWh'
    unidade_medida = 'm³' if utilidade == 'Água' else 'kWh'
    cor_fin = '#42a5c8' if utilidade == 'Água' else '#e67e3a'
    cor_fis = '#5ec8e8' if utilidade == 'Água' else '#f4b24a'

    c1, c2 = st.columns(2)

    with c1:
        rank_fin = df.groupby('Unidade')[['Valor_Total', col_consumo]].sum().reset_index().sort_values('Valor_Total', ascending=False).head(15)
        rank_fin['Valor_Fmt'] = rank_fin['Valor_Total'].apply(formatar_moeda_br)
        rank_fin['Consumo_Fmt'] = rank_fin[col_consumo].apply(formatar_numero_br)
        fig_fin = px.bar(rank_fin, x='Unidade', y='Valor_Total', custom_data=['Valor_Fmt', 'Consumo_Fmt'], title="Maior custo financeiro")
        fig_fin.update_traces(marker_color=cor_fin, hovertemplate=f"<b>Unidade:</b> %{{x}}<br><b>Custo:</b> %{{customdata[0]}}<br><b>Consumo:</b> %{{customdata[1]}} {unidade_medida}<extra></extra>")
        fig_fin.update_layout(xaxis={'categoryorder': 'total descending'}, yaxis_tickformat=",.0f", xaxis_title="", yaxis_title="R$")
        fig_fin = plotly_dark(fig_fin)
        st.plotly_chart(fig_fin, use_container_width=True)

    with c2:
        rank_fis = df.groupby('Unidade')[['Valor_Total', col_consumo]].sum().reset_index().sort_values(col_consumo, ascending=False).head(15)
        rank_fis['Valor_Fmt'] = rank_fis['Valor_Total'].apply(formatar_moeda_br)
        rank_fis['Consumo_Fmt'] = rank_fis[col_consumo].apply(formatar_numero_br)
        fig_fis = px.bar(rank_fis, x='Unidade', y=col_consumo, custom_data=['Valor_Fmt', 'Consumo_Fmt'], title="Maior consumo físico")
        fig_fis.update_traces(marker_color=cor_fis, hovertemplate=f"<b>Unidade:</b> %{{x}}<br><b>Custo:</b> %{{customdata[0]}}<br><b>Consumo:</b> %{{customdata[1]}} {unidade_medida}<extra></extra>")
        fig_fis.update_layout(xaxis={'categoryorder': 'total descending'}, yaxis_tickformat=",.0f", xaxis_title="", yaxis_title=unidade_medida)
        fig_fis = plotly_dark(fig_fis)
        st.plotly_chart(fig_fis, use_container_width=True)

def render_ai_assistant(df_filtrado, utilidade):
    slug = utilidade.replace(" ", "_")
    st.markdown(f"#### Perguntas rápidas · {utilidade}")
    st.caption("Respostas calculadas na hora com os mesmos filtros dos cartões acima.")

    perguntas = [
        ("q1", "Quem lidera em custo?"),
        ("q2", "Quem lidera em consumo?"),
        ("q3", "Qual mês teve maior fatura?"),
        ("q4", "Onde está o maior volume de gasto (origem)?"),
        ("q5", "Top 3 unidades por custo"),
        ("q6", "Onde a tarifa média mais pesa?"),
        ("q7", "Onde houve a maior alta no custo (mês a mês)?"),
        ("q8", "Quanto são encargos vs faturamento?"),
    ]

    ncols = 4
    for row_start in range(0, len(perguntas), ncols):
        row = perguntas[row_start : row_start + ncols]
        cols = st.columns(ncols)
        for col, (qid, label) in zip(cols, row):
            with col:
                if st.button(label, key=f"quick_{slug}_{qid}", use_container_width=True):
                    st.session_state[f"sel_q_{slug}"] = qid

    sel_key = f"sel_q_{slug}"
    if sel_key in st.session_state and st.session_state[sel_key]:
        qid = st.session_state[sel_key]
        texto = resposta_pergunta_rapida(df_filtrado, utilidade, qid)
        render_caixa_resposta(texto)

    with st.expander("Pergunta livre com IA (Gemini)", expanded=False):
        st.caption(
            "Perguntas abertas usam a IA (Gemini). As respostas rápidas são calculadas na hora a partir dos dados "
            "carregados — sem IA. Quando a planilha é atualizada e os dados são lidos de novo, esses resultados mudam automaticamente."
        )
        col_ia_input, col_ia_btn = st.columns([5, 1])
        with col_ia_input:
            pergunta_ia = st.text_input(
                "Pergunta",
                key=f"ia_input_{slug}",
                label_visibility="collapsed",
                placeholder="Ex.: Compare o custo entre duas unidades em meses específicos…",
            )
        with col_ia_btn:
            btn_consultar = st.button("Enviar", key=f"btn_{slug}", use_container_width=True, type="primary")

        if btn_consultar:
            key = get_gemini_key()
            if not key:
                st.error("Defina GEMINI_API_KEY nas secrets.")
            elif not pergunta_ia:
                st.warning("Digite uma pergunta.")
            else:
                with st.spinner("Consultando Gemini…"):
                    try:
                        genai.configure(api_key=key)
                        modelos_permitidos = [
                            m.name for m in genai.list_models() if "generateContent" in m.supported_generation_methods
                        ]
                        if not modelos_permitidos:
                            st.error("Chave sem modelos de texto liberados.")
                            st.stop()

                        modelo_ideal = next((m for m in modelos_permitidos if "1.5-flash" in m), modelos_permitidos[0])
                        nome_limpo = modelo_ideal.replace("models/", "")
                        modelo = genai.GenerativeModel(nome_limpo)

                        palavras_chave_cadastrais = [
                            "UNIDADE EXECUTORA",
                            "UPG",
                            "HIDRÔMETRO",
                            "CONCESSIONÁRIA",
                            "CNPJ",
                            "IDENTIFICADOR",
                            "EMPENHO",
                            "MATRÍCULA",
                            "FONTE 10",
                            "FONTE 60",
                        ]
                        cols_valor = ["Valor_Total", "Valor_Encargos"]
                        cols_m3 = ["Volume_M3"] if utilidade == "Água" else ["Consumo_KWh"]

                        colunas_cadastrais_encontradas = [
                            col
                            for col in df_filtrado.columns
                            if any(pc in str(col).upper() for pc in palavras_chave_cadastrais)
                            and col not in cols_valor + cols_m3 + ["Unidade", "Nome_Mes"]
                        ]

                        if cols_valor or cols_m3:
                            agg_dict = {c: "sum" for c in cols_valor + cols_m3 if c in df_filtrado.columns}
                            agg_dict.update({c: "first" for c in colunas_cadastrais_encontradas})
                            df_ia = df_filtrado.groupby(["Unidade", "Nome_Mes"]).agg(agg_dict).reset_index()
                            contexto_dados = df_ia.to_csv(index=False, sep="\t")
                        else:
                            contexto_dados = "Sem dados financeiros na planilha filtrada."

                        prompt_ia = f"""
                        Você é um analista de dados da SEJUSP.
                        Dados de {utilidade.upper()} (unidades prisionais e socioeducativas), faturamento mês a mês.

                        REGRAS:
                        - Nome_Mes: mês de referência (Jan, Fev…).
                        - Valor_Total: custo em R$.
                        - Volume_M3 ou Consumo_KWh: consumo físico.

                        DADOS:
                        {contexto_dados}

                        PERGUNTA: {pergunta_ia}

                        Responda com base só na tabela. Valores em R$ quando aplicável.
                        Use texto corrido em parágrafos curtos. Não use markdown (sem asteriscos, sem negrito, sem listas com hífen no início da linha).
                        """
                        resposta = modelo.generate_content(prompt_ia)
                        render_caixa_resposta(resposta.text)
                    except Exception as e:
                        st.error(f"Erro na IA: {e}")

def render_manual():
    st.markdown("**Roteiro DSG / SEJUSP**")
    st.markdown("""
    **Gestão de Faturas de Utilidade Pública (Água e Energia)**
    
    #### 1. Atualização das Planilhas
    * O lançamento das faturas é realizado nas abas correspondentes à Concessionária (ex: COPASA, CESAMA, Diversas).
    * As faturas chegam por e-mail: a equipe monta o processo no SEI e lança o número e dados da fatura na planilha.

    #### 2. Códigos de Unidade Executora (UE)
    * **SAÚDE:** `1450005` *(Ordenador: Ana Louise)*
    * **SUASE:** `1450012` *(Ordenador: Giselle)*
    * **SUINT:** `1450015` *(Ordenador: Christian)*
    * **DEPEN:** `1450022` *(Ordenador: Carlos)*

    #### 3. Emissão e Consulta de Empenhos (SIAFI)
    * **Criar Empenho:** Utilize o atalho `F5 + F5` e anote o número do empenho criado antes de pressionar ENTER.
    * **Consultar Empenho:** Acesse a Navegação `95/4` (Empenhos Emitidos) > `1` (Consultar) > Marque credor e insira o CNPJ da Concessionária.
    * Após finalizar a criação, não se esqueça de emitir a nota de empenho no SIAFI.
    * **Conferência:** Conferir o saldo. Se for insuficiente, providenciar reforço ou descentralização.

    #### 4. Códigos de Descentralização (SULOT / Outros Setores)
    *(Caso falte saldo, solicitar descentralização por Natureza de Despesa)*
    * **SULOT:** `2500`
    * **SUPEC:** `4344`
    * **SUASE:** `4441`
    * **SUINT:** `4378`
    * **DEPEN:** `4348`
    * **SAÚDE:** `4353`

    #### 5. Cancelamento e Restos a Pagar
    * Para realizar o cancelamento de restos a pagar não processados, utilize a **Navegação: 51** no SIAFI.
    """)


# --- 5. LÓGICA DA INTERFACE PRINCIPAL ---
caminho_agua = _secret("CAMINHO_AGUA") or os.getenv("CAMINHO_AGUA", "")
caminho_energia = _secret("CAMINHO_ENERGIA") or os.getenv("CAMINHO_ENERGIA", "")

with st.sidebar:
    st.markdown("### Fonte dos dados")
    st.caption(
        "Configure **CAMINHO_AGUA** e **CAMINHO_ENERGIA** em `.streamlit/secrets.toml` "
        "(local) ou nas secrets do Streamlit Cloud. Use caminho de pasta sincronizada ou URL de download."
    )
    f_agua = st.file_uploader("Substituir Água (.xlsx)", type=["xlsx"], key="up_agua")
    f_energia = st.file_uploader("Substituir Energia (.xlsx)", type=["xlsx"], key="up_energia")
    if not caminho_agua and f_agua is None:
        st.info("Água: defina o caminho/URL nas secrets ou envie um arquivo.")
    if not caminho_energia and f_energia is None:
        st.info("Energia: defina o caminho/URL nas secrets ou envie um arquivo.")
    if not get_gemini_key():
        st.warning("Assistente IA: defina **GEMINI_API_KEY** nas secrets.")

if f_agua is not None:
    df_agua_raw = carregar_dados_upload(f_agua, "agua")
else:
    df_agua_raw = carregar_dados_cache(caminho_agua, "agua") if caminho_agua else pd.DataFrame()

df_agua = enriquecer_dados(df_agua_raw, "agua")

if f_energia is not None:
    df_energia_raw = carregar_dados_upload(f_energia, "energia")
else:
    df_energia_raw = carregar_dados_cache(caminho_energia, "energia") if caminho_energia else pd.DataFrame()

df_energia = enriquecer_dados(df_energia_raw, "energia")

st.markdown(
    """
<div class="hero-banner">
    <div class="hero-kicker">SEJUSP · DSG</div>
    <h1 class="hero-main-title">DASHBOARD NNUP</h1>
    <p class="hero-sub">Controle financeiro de Água e Energia da SEJUSP MG.</p>
    <div class="hero-accent"></div>
</div>
""",
    unsafe_allow_html=True,
)

aba_agua, aba_energia = st.tabs(["ÁGUA", "ENERGIA"])

# ==========================================
# ABA 1: ÁGUA
# ==========================================
with aba_agua:
    if df_agua.empty:
        st.warning("Dados de Água não encontrados. Verifique se a planilha não está aberta no Excel.")
    else:
        menu_agua = st.radio("Visões disponíveis:", 
                             ["GERAL (Ranking)", "DEPEN", "SUASE", "SUINT", "MANUAL", "ALERTA"], 
                             horizontal=True, label_visibility="collapsed")
        
        if menu_agua == "MANUAL":
            render_manual()
        else:
            st.markdown("---")
            col_busca_a1, col_busca_a2 = st.columns(2)
            unidades_agua = ["Todas"] + sorted(df_agua['Unidade'].astype(str).unique())
            
            # Filtro Unidade e Filtro Mês lado a lado
            busca_unidade_agua = col_busca_a1.selectbox("🔎 Filtrar por Unidade (Água):", unidades_agua)
            busca_mes_agua = col_busca_a2.selectbox("📅 Filtrar por Mês:", LISTA_MESES, key="mes_agua")
            
            dff_a = df_agua.copy()
            if busca_unidade_agua != "Todas":
                dff_a = dff_a[dff_a['Unidade'] == busca_unidade_agua]
            if busca_mes_agua != "Todos":
                mes_filtrado_a = MAPA_MESES[busca_mes_agua]
                dff_a = dff_a[dff_a['Nome_Mes'] == mes_filtrado_a]

            if menu_agua == "DEPEN":
                dff_a = dff_a[dff_a['Origem'].str.contains('GLOBALIZADA COPASA|GLOBALIZADA CESAMA|DIVERSAS', case=False, na=False)]
            elif menu_agua == "SUASE":
                dff_a = dff_a[dff_a['Origem'].str.contains('SUASE', case=False, na=False)]
            elif menu_agua == "SUINT":
                dff_a = dff_a[dff_a['Origem'].str.contains('SUINT', case=False, na=False)]

            gasto_total_a = dff_a['Valor_Total'].sum()
            consumo_total_a = dff_a['Volume_M3'].sum()
            media_a = gasto_total_a / consumo_total_a if consumo_total_a > 0 else 0

            c1, c2, c3 = st.columns(3)
            c1.metric("💸 Custo Total Faturado", formatar_moeda_br(gasto_total_a))
            c2.metric("💧 Volume Total Consumido", f"{formatar_numero_br(consumo_total_a)} m³")
            c3.metric("📊 Tarifa Média (R$/m³)", f"{formatar_moeda_br(media_a)}")

            st.markdown("<br>", unsafe_allow_html=True)
            
            if menu_agua != "ALERTA":
                render_ai_assistant(dff_a, "Água")

            if menu_agua in ["GERAL (Ranking)", "DEPEN", "SUASE", "SUINT"]:
                if menu_agua == "GERAL (Ranking)":
                    st.markdown("**Visão geral**")
                    col_graf_a1, col_graf_a2 = st.columns(2)
                    with col_graf_a1:
                        donut_a = dff_a.groupby('Origem')['Valor_Total'].sum().reset_index()
                        donut_a['Valor_Fmt'] = donut_a['Valor_Total'].apply(formatar_moeda_br)
                        fig_donut_a = px.pie(donut_a, names='Origem', values='Valor_Total', hole=0.45, custom_data=['Valor_Fmt'], title="Orçamento por origem")
                        fig_donut_a.update_traces(hovertemplate="<b>%{label}</b><br>Custo: %{customdata[0]}<br>Representatividade: %{percent}<extra></extra>")
                        fig_donut_a = plotly_dark(fig_donut_a)
                        st.plotly_chart(fig_donut_a, use_container_width=True)

                    with col_graf_a2:
                        evo_global_a = dff_a.groupby(['Mes', 'Nome_Mes'])['Valor_Total'].sum().reset_index().sort_values('Mes')
                        evo_global_a['Valor_Fmt'] = evo_global_a['Valor_Total'].apply(formatar_moeda_br)
                        fig_evo_global_a = px.bar(evo_global_a, x='Nome_Mes', y='Valor_Total', custom_data=['Valor_Fmt'], title="Faturamento por mês")
                        fig_evo_global_a.update_traces(marker_color="#42a5c8", hovertemplate="<b>Mês:</b> %{x}<br><b>Custo Total:</b> %{customdata[0]}<extra></extra>")
                        fig_evo_global_a.update_layout(xaxis_title="", yaxis_title="R$", yaxis_tickformat=",.0f")
                        fig_evo_global_a = plotly_dark(fig_evo_global_a)
                        st.plotly_chart(fig_evo_global_a, use_container_width=True)

                elif menu_agua in ["DEPEN", "SUASE", "SUINT"]:
                    st.markdown(f"**Setor · {menu_agua}**")
                    evo_a = dff_a.groupby(['Mes', 'Nome_Mes'])[['Valor_Total', 'Volume_M3']].sum().reset_index().sort_values('Mes')
                    evo_a['Valor_Total_Fmt'] = evo_a['Valor_Total'].apply(formatar_moeda_br)
                    evo_a['Consumo_Fmt'] = evo_a['Volume_M3'].apply(formatar_numero_br)
                    fig_evo_a = px.bar(evo_a, x='Nome_Mes', y='Valor_Total', custom_data=['Valor_Total_Fmt', 'Consumo_Fmt'], title="Faturamento mensal")
                    fig_evo_a.update_traces(marker_color="#5ec8e8", hovertemplate="<b>Mês:</b> %{x}<br><b>Valor:</b> %{customdata[0]}<br><b>Volume:</b> %{customdata[1]} m³<extra></extra>")
                    fig_evo_a.update_layout(xaxis_title="", yaxis_title="", yaxis_tickformat=",.0f")
                    fig_evo_a = plotly_dark(fig_evo_a)
                    st.plotly_chart(fig_evo_a, use_container_width=True)

                    st.markdown("---")
                    st.markdown("**Tarifa vs volume**")
                    st.caption("Bolhas altas: tarifa acima da média; à direita: consumo elevado.")
                    scatter_df_a = dff_a.groupby('Unidade')[['Valor_Total', 'Volume_M3']].sum().reset_index()
                    scatter_df_a = scatter_df_a[scatter_df_a['Volume_M3'] > 0]
                    scatter_df_a['Preco_Medio'] = scatter_df_a['Valor_Total'] / scatter_df_a['Volume_M3']
                    scatter_df_a['Valor_Fmt'] = scatter_df_a['Valor_Total'].apply(formatar_moeda_br)
                    scatter_df_a['Volume_Fmt'] = scatter_df_a['Volume_M3'].apply(formatar_numero_br)
                    scatter_df_a['Preco_Fmt'] = scatter_df_a['Preco_Medio'].apply(formatar_moeda_br)
                    fig_scatter_a = px.scatter(scatter_df_a, x='Volume_M3', y='Preco_Medio', size='Valor_Total', color='Preco_Medio', custom_data=['Valor_Fmt', 'Volume_Fmt', 'Preco_Fmt', 'Unidade'], color_continuous_scale='Viridis')
                    fig_scatter_a.update_traces(hovertemplate="<b>%{customdata[3]}</b><br>Volume: %{customdata[1]} m³<br>Tarifa Média: %{customdata[2]}/m³<br>Custo Total: %{customdata[0]}<extra></extra>")
                    fig_scatter_a.update_layout(xaxis_title="Volume (m³)", yaxis_title="R$/m³")
                    fig_scatter_a = plotly_dark(fig_scatter_a)
                    st.plotly_chart(fig_scatter_a, use_container_width=True)
                
                exibir_top_15(dff_a, "Água")

            elif menu_agua == "ALERTA":
                st.markdown("**Alertas · água**")
                col_alerta_m3, col_alerta_fin = st.columns(2)
                with col_alerta_m3:
                    st.markdown('<div class="danger-box"><b>🔥 Explosão de Consumo (>20%)</b></div>', unsafe_allow_html=True)
                    alert_m3 = dff_a[(dff_a['Var_Consumo_Pct']>20) & (dff_a['Volume_M3']>50)].sort_values('Var_Consumo_Pct', ascending=False)
                    if not alert_m3.empty: st.dataframe(alert_m3[['Unidade', 'Nome_Mes', 'Volume_M3', 'Var_Consumo_Pct']], hide_index=True, use_container_width=True)
                    else: st.success("Nenhuma anomalia de consumo.")
                with col_alerta_fin:
                    st.markdown('<div class="money-box"><b>💸 Explosão Financeira (>20%)</b></div>', unsafe_allow_html=True)
                    alert_fin = dff_a[(dff_a['Var_Financeiro_Pct']>20) & (dff_a['Valor_Total']>500)].sort_values('Var_Financeiro_Pct', ascending=False)
                    if not alert_fin.empty: st.dataframe(alert_fin[['Unidade', 'Nome_Mes', 'Valor_Total', 'Var_Financeiro_Pct']], hide_index=True, use_container_width=True)
                    else: st.success("Nenhuma anomalia financeira.")


# ==========================================
# ABA 2: ENERGIA
# ==========================================
with aba_energia:
    if df_energia.empty:
        st.warning("Dados de Energia não encontrados. Verifique se a planilha não está aberta no Excel.")
    else:
        menu_energia = st.radio("Visões disponíveis - ENERGIA:", 
                             ["GERAL (Ranking)", "BAIXA TENSÃO", "MÉDIA TENSÃO", "MANUAL", "ALERTA"], 
                             horizontal=True, label_visibility="collapsed")
        
        if menu_energia == "MANUAL":
            render_manual()
        else:
            st.markdown("---")
            col_busca_e1, col_busca_e2 = st.columns(2)
            unidades_energia = ["Todas"] + sorted(df_energia['Unidade'].astype(str).unique())
            
            # Filtro Unidade e Filtro Mês lado a lado
            busca_unidade_energia = col_busca_e1.selectbox("🔎 Filtrar por Unidade (Energia):", unidades_energia)
            busca_mes_energia = col_busca_e2.selectbox("📅 Filtrar por Mês:", LISTA_MESES, key="mes_energia")
            
            dff_e = df_energia.copy()
            if busca_unidade_energia != "Todas":
                dff_e = dff_e[dff_e['Unidade'] == busca_unidade_energia]
            if busca_mes_energia != "Todos":
                mes_filtrado_e = MAPA_MESES[busca_mes_energia]
                dff_e = dff_e[dff_e['Nome_Mes'] == mes_filtrado_e]

            if menu_energia in ["BAIXA TENSÃO", "MÉDIA TENSÃO"]:
                sub_setor = st.selectbox("Setor:", ["Todos", "DEPEN", "SUASE", "SUINT"])
                if sub_setor != "Todos":
                    dff_e = dff_e[dff_e['Origem'].str.contains(sub_setor, case=False, na=False)]

            gasto_total_e = dff_e['Valor_Total'].sum()
            consumo_total_e = dff_e['Consumo_KWh'].sum()
            media_e = gasto_total_e / consumo_total_e if consumo_total_e > 0 else 0

            c1, c2, c3 = st.columns(3)
            c1.metric("⚡ Custo Total Faturado", formatar_moeda_br(gasto_total_e))
            c2.metric("🔋 Demanda Acumulada", f"{formatar_numero_br(consumo_total_e)} kWh")
            c3.metric("📈 Tarifa Média (R$/kWh)", f"{formatar_moeda_br(media_e)}")

            st.markdown("<br>", unsafe_allow_html=True)
            
            if menu_energia != "ALERTA":
                render_ai_assistant(dff_e, "Energia")

            if menu_energia in ["GERAL (Ranking)", "BAIXA TENSÃO", "MÉDIA TENSÃO"]:
                if menu_energia == "GERAL (Ranking)":
                    st.markdown("**Visão geral**")
                    col_graf_e1, col_graf_e2 = st.columns(2)
                    with col_graf_e1:
                        donut_e = dff_e.groupby('Origem')['Valor_Total'].sum().reset_index()
                        donut_e['Valor_Fmt'] = donut_e['Valor_Total'].apply(formatar_moeda_br)
                        fig_donut_e = px.pie(donut_e, names='Origem', values='Valor_Total', hole=0.45, custom_data=['Valor_Fmt'], title="Orçamento por origem")
                        fig_donut_e.update_traces(hovertemplate="<b>%{label}</b><br>Custo: %{customdata[0]}<br>Representatividade: %{percent}<extra></extra>")
                        fig_donut_e = plotly_dark(fig_donut_e)
                        st.plotly_chart(fig_donut_e, use_container_width=True)

                    with col_graf_e2:
                        evo_global_e = dff_e.groupby(['Mes', 'Nome_Mes'])['Valor_Total'].sum().reset_index().sort_values('Mes')
                        evo_global_e['Valor_Fmt'] = evo_global_e['Valor_Total'].apply(formatar_moeda_br)
                        fig_evo_global_e = px.bar(evo_global_e, x='Nome_Mes', y='Valor_Total', custom_data=['Valor_Fmt'], title="Faturamento por mês")
                        fig_evo_global_e.update_traces(marker_color="#e67e3a", hovertemplate="<b>Mês:</b> %{x}<br><b>Custo Total:</b> %{customdata[0]}<extra></extra>")
                        fig_evo_global_e.update_layout(xaxis_title="", yaxis_title="R$", yaxis_tickformat=",.0f")
                        fig_evo_global_e = plotly_dark(fig_evo_global_e)
                        st.plotly_chart(fig_evo_global_e, use_container_width=True)

                elif menu_energia in ["BAIXA TENSÃO", "MÉDIA TENSÃO"]:
                    st.markdown(f"**{menu_energia}**")
                    evo_e = dff_e.groupby(['Mes', 'Nome_Mes'])[['Valor_Total', 'Consumo_KWh']].sum().reset_index().sort_values('Mes')
                    evo_e['Valor_Total_Fmt'] = evo_e['Valor_Total'].apply(formatar_moeda_br)
                    evo_e['Consumo_Fmt'] = evo_e['Consumo_KWh'].apply(formatar_numero_br)
                    fig_evo_e = px.bar(evo_e, x='Nome_Mes', y='Valor_Total', custom_data=['Valor_Total_Fmt', 'Consumo_Fmt'], title="Faturamento mensal")
                    fig_evo_e.update_traces(marker_color="#f4b24a", hovertemplate="<b>Mês:</b> %{x}<br><b>Valor:</b> %{customdata[0]}<br><b>Demanda:</b> %{customdata[1]} kWh<extra></extra>")
                    fig_evo_e.update_layout(xaxis_title="", yaxis_title="", yaxis_tickformat=",.0f")
                    fig_evo_e = plotly_dark(fig_evo_e)
                    st.plotly_chart(fig_evo_e, use_container_width=True)

                    st.markdown("---")
                    st.markdown("**Tarifa vs demanda**")
                    st.caption("Topo: custo por kWh alto; à direita: demanda elevada.")
                    scatter_df_e = dff_e.groupby('Unidade')[['Valor_Total', 'Consumo_KWh']].sum().reset_index()
                    scatter_df_e = scatter_df_e[scatter_df_e['Consumo_KWh'] > 0]
                    scatter_df_e['Preco_Medio'] = scatter_df_e['Valor_Total'] / scatter_df_e['Consumo_KWh']
                    scatter_df_e['Valor_Fmt'] = scatter_df_e['Valor_Total'].apply(formatar_moeda_br)
                    scatter_df_e['Consumo_Fmt'] = scatter_df_e['Consumo_KWh'].apply(formatar_numero_br)
                    scatter_df_e['Preco_Fmt'] = scatter_df_e['Preco_Medio'].apply(formatar_moeda_br)
                    fig_scatter_e = px.scatter(scatter_df_e, x='Consumo_KWh', y='Preco_Medio', size='Valor_Total', color='Preco_Medio', custom_data=['Valor_Fmt', 'Consumo_Fmt', 'Preco_Fmt', 'Unidade'], color_continuous_scale='OrRd')
                    fig_scatter_e.update_traces(hovertemplate="<b>%{customdata[3]}</b><br>Demanda: %{customdata[1]} kWh<br>Tarifa: %{customdata[2]}/kWh<br>Custo Total: %{customdata[0]}<extra></extra>")
                    fig_scatter_e.update_layout(xaxis_title="kWh", yaxis_title="R$/kWh")
                    fig_scatter_e = plotly_dark(fig_scatter_e)
                    st.plotly_chart(fig_scatter_e, use_container_width=True)
                
                exibir_top_15(dff_e, "Energia")

            elif menu_energia == "ALERTA":
                st.markdown("**Alertas · energia**")
                col_alerta_e_kwh, col_alerta_e_fin = st.columns(2)
                with col_alerta_e_kwh:
                    st.markdown('<div class="warning-box"><b>⚠️ Pico de Demanda (>15%)</b></div>', unsafe_allow_html=True)
                    alert_kwh = dff_e[(dff_e['Var_Consumo_Pct']>15) & (dff_e['Consumo_KWh']>100)].sort_values('Var_Consumo_Pct', ascending=False)
                    if not alert_kwh.empty: st.dataframe(alert_kwh[['Unidade', 'Nome_Mes', 'Consumo_KWh', 'Var_Consumo_Pct']], hide_index=True, use_container_width=True)
                    else: st.success("Consumo energético normal.")
                with col_alerta_e_fin:
                    st.markdown('<div class="money-box"><b>💸 Explosão Financeira (>15%)</b></div>', unsafe_allow_html=True)
                    alert_fin_e = dff_e[(dff_e['Var_Financeiro_Pct']>15) & (dff_e['Valor_Total']>500)].sort_values('Var_Financeiro_Pct', ascending=False)
                    if not alert_fin_e.empty: st.dataframe(alert_fin_e[['Unidade', 'Nome_Mes', 'Valor_Total', 'Var_Financeiro_Pct']], hide_index=True, use_container_width=True)
                    else: st.success("Financeiro energético normal.")